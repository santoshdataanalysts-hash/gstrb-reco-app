import streamlit as st
import pandas as pd

from io import BytesIO
from utils import *
import utils

# -----------------------------------
# PAGE CONFIG
# -----------------------------------

st.set_page_config(
    page_title="GST Reconciliation",
    layout="wide"
)

st.title("GST 2B Reconciliation")


# -----------------------------------
# FILE UPLOAD
# -----------------------------------

purchase_file = st.file_uploader(
    "Upload Purchase Register",
    type=["xlsx"]
)

gstr2b_file = st.file_uploader(
    "Upload GSTR2B",
    type=["xlsx"]
)

# -----------------------------------
# PROCESS
# -----------------------------------
if purchase_file and gstr2b_file:

    st.success("Files Uploaded Successfully")

    # READ FILES
    purchase_df = utils.read_purchase_file(
        purchase_file
    )
    gstr2b_df = utils.read_gstr2b_file(
        gstr2b_file
    )
    purchase_df = utils.standardize_columns(
        purchase_df,
        "purchase"
    )

    purchase_df = utils.create_total_tax(
        purchase_df
    )
    # FIX BLANK DETAIL ROWS

    fill_cols = [
        "GSTIN",
        "PARTY_NAME",
        "INVOICE_NO",
        "INVOICE_DATE"
    ]

    for col in fill_cols:
        if col in purchase_df.columns:
            purchase_df[col] = purchase_df[col].ffill()
    
    gstr2b_df = utils.standardize_columns(
        gstr2b_df,
        "gstr2b"
    )
    # TOTAL TAX
    purchase_df = utils.create_total_tax(
        purchase_df
    )

    gstr2b_df = utils.create_total_tax(
        gstr2b_df
    )
    # -----------------------------------
    # REQUIRED COLUMN CHECK
    # ----------------------------------

    required_purchase_cols = [
        "GSTIN",
        "INVOICE_NO",
        "TAXABLE_VALUE_REC"
    ]

    missing_purchase = [
        col
        for col in required_purchase_cols
        if col not in purchase_df.columns
    ]

    if missing_purchase:

        st.error(
            f"Missing Purchase Columns: {missing_purchase}"
        )

        st.stop()

    required_2b_cols = [
        "GSTIN",
        "INVOICE_NO",
        "TAXABLE_VALUE"
    ]

    missing_2b = [
        col
        for col in required_2b_cols
        if col not in gstr2b_df.columns
    ]

    if missing_2b:

        st.error(
            f"Missing GSTR2B Columns: {missing_2b}"
        )

        st.stop()
    
    # SAFE INVOICE COLUMN
    if "INVOICE_NO" not in purchase_df.columns:

        st.error("Purchase File Invoice Column Not Found")
        st.stop()

    if "INVOICE_NO" not in gstr2b_df.columns:

        st.error("GSTR2B Invoice Column Not Found")
        st.stop()

    # CLEAN INVOICE
    purchase_df["normalize_invoice"] = (
        purchase_df["INVOICE_NO"]
        .astype(str)
        .apply(utils.normalize_invoice)
    )

    gstr2b_df["normalize_invoice"] = (
        gstr2b_df["INVOICE_NO"]
        .astype(str)
        .apply(utils.normalize_invoice)
    )
    # RECON
    final_df, cdnr_df, extra_2b = reconcile_data(
        purchase_df,
        gstr2b_df
    )
    # CHECK TOP MISSING GSTIN

    missing_df = final_df[
        final_df["RECON_STATUS"] == "MISSING IN 2B"
    ]

    top_vendors = (
        missing_df.groupby("GSTIN")
        .size()
        .sort_values(ascending=False)
        .head(20)
    )

    if final_df.empty:

        st.warning("No Data Found After Reconciliation")
        st.stop()
    
    # FIX OBJECT CONVERSION ERROR

    object_cols = final_df.select_dtypes(
        include=["object"]
    ).columns

    for col in object_cols:

        final_df[col] = (
            final_df[col]
            .fillna("")
            .astype(str)
        )
    
    # SHOW RESULT
    st.write("FINAL RECONCILIATION")
    
    st.write(final_df)

    # DOWNLOAD
    output = BytesIO()

    download_df = final_df.copy()
    required_columns = [

        "TRAN_TYPE",
        "GSTIN",
        "PARTY_NAME",
        "INVOICE_DATE",
        "INVOICE_NO",
        "BILL_AMOUNT",
        "INVOICE_TYPE",
        "STATE_CODE",
        "TAXABLE_VALUE_REC",
        "TAXABLE_VALUE_PAY",
        "IGST",
        "CGST",
        "SGST",
        "DIVISION",
        "TOTAL_TAX",
        "TAXABLE_VALUE_2B",
        "TOTAL_TAX_2B",
        "RECON_STATUS",
        "REMARK"
    ]

    download_df = download_df[
        [
            col for col in required_columns
            if col in download_df.columns
        ]
    ]
    # SAFE STRING CONVERSION

    object_cols = download_df.select_dtypes(
        include=["object"]
    ).columns

    for col in object_cols:

        download_df[col] = (
            download_df[col]
            .fillna("")
            .astype(str)
        )
    #KEEP AVAILABLE COLUMNS ONLY
    final_columns = [

        col
        for col in required_columns
        if col in download_df.columns
    ]

    download_df = download_df[
        final_columns
    ]
    # SORT BY STATUS
    if "RECON_STATUS" in download_df.columns:

        download_df = download_df.sort_values(
            by="RECON_STATUS"
        )
    # RESET INDEX
    download_df = download_df.reset_index(
        drop=True
    )

    # INVOICE FORMAT
    if "INVOICE_NO" in download_df.columns:
        download_df["INVOICE_NO"] = (
            download_df["INVOICE_NO"]
            .astype(str)
        )
    
    #CLEAN COLUMNS NAMES
    download_df.columns = (
        download_df.columns
        .str.strip()
    )
    # REMOVE EXTRA COLUMNS
    remove_cols = [
        "LEDGER AMT (WITH ROUNDOFF)",
        "SALE/PURCH A/C",
        "REF INVOICE",
        "REF INV DATE",
        "normalize_invoice",
        "CESS",
        "MATCH_KEY"
    ]

    download_df = download_df.drop(
        columns=[
            col
            for col in remove_cols
            if col in download_df.columns
        ],
        errors="ignore"
    )

    # REMOVE DUPLICATE
    download_df = download_df.loc[
        :,
        ~download_df.columns.duplicated()
    ]
    cn_dn_df = gstr2b_df[
        gstr2b_df["SOURCE_SHEET"].isin(
            ["B2B-CDNR", "B2B-CDNRA"]
        )
    ].copy()
    # -----------------------------------
    # VENDOR SUMMARY REPORT
    # -----------------------------------

    gstr2b_df["TAXABLE_VALUE"] = pd.to_numeric(
        gstr2b_df["TAXABLE_VALUE"],
        errors="coerce"
    ).fillna(0)

    purchase_df["TAXABLE_VALUE_REC"] = pd.to_numeric(
        purchase_df["TAXABLE_VALUE_REC"],
        errors="coerce"
    ).fillna(0)

    vendor_2b = (
        gstr2b_df.groupby(["GSTIN", "PARTY_NAME"])
        .agg(
            BILLS_IN_2B=("INVOICE_NO", "nunique"),
            TAXABLE_AMT_2B=("TAXABLE_VALUE", "sum")
        )
        .reset_index()
    )

    vendor_pr = (
        purchase_df.groupby(["GSTIN", "PARTY_NAME"])
        .agg(
            BILLS_IN_PURCHASE=("INVOICE_NO", "nunique"),
            TAXABLE_AMT_PURCHASE=("TAXABLE_VALUE_REC", "sum")
        )
        .reset_index()
    )

    vendor_summary = pd.merge(
        vendor_2b,
        vendor_pr,
        on=["GSTIN", "PARTY_NAME"],
        how="outer"
    ).fillna(0)

    vendor_summary["DIFFERENCE"] = (
        vendor_summary["TAXABLE_AMT_2B"]
        - vendor_summary["TAXABLE_AMT_PURCHASE"]
    )

    vendor_summary = vendor_summary.sort_values(
        by="DIFFERENCE",
        ascending=False
    )
    
    from io import BytesIO
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    b2b_df = download_df.copy()
    cdnr_df = cdnr_df.copy()
    extra_2b = extra_2b.copy()
    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        download_df.to_excel(
            writer,
            sheet_name="RECO_REPORT",
            index=False
        )

        cdnr_df.to_excel(
            writer,
            sheet_name="CN_DN_IN_2B",
            index=False
        )
        extra_2b.to_excel(
            writer,
            sheet_name="MISSING IN PURCHASE",
            index=False
        )
        vendor_summary.to_excel(
            writer,
            sheet_name="VENDOR_SUMMARY",
            index=False
        )       
        for sheet_name in ["RECO_REPORT", "CN_DN_IN_2B"]:

            ws = writer.sheets[sheet_name]

            # RECON_STATUS Column Find
            status_col = None

            for cell in ws[1]:

                if cell.value == "RECON_STATUS":

                    status_col = cell.column
                    break

            # Status Color Formatting
            if status_col:

                for row in ws.iter_rows(min_row=2):

                    status_cell = row[status_col - 1]

                    status = status_cell.value

                    if status == "MATCHED":

                        status_cell.font = Font(
                            color="008000",
                            bold=True
                        )

                    elif status == "MISSING IN 2B":

                        status_cell.font = Font(
                            color="FF0000",
                            bold=True
                        )

                    elif status == "MISSING IN PURCHASE":

                        status_cell.font = Font(
                            color="FF8C00",
                            bold=True
                        )

                    elif status == "TAXABLE VALUE MISMATCH":

                        status_cell.font = Font(
                            color="B8860B",
                            bold=True
                        )

                    elif status == "TAX AMOUNT DIFFERENCE":

                        status_cell.font = Font(
                            color="800080",
                            bold=True
                        )

            # Filter + Freeze
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            # Auto Width
            for column_cells in ws.columns:

                max_length = 0

                for cell in column_cells:

                    try:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                    except:
                        pass

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                ws.column_dimensions[
                    column_letter
                ].width = min(max_length + 2, 40)

    output.seek(0)

    st.download_button(
        label="Download Reco Report",
        data=output.getvalue(),
        file_name="GST_RECON_REPORT.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
